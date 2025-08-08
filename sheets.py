from dataclasses import dataclass
from datetime import datetime
from itertools import permutations
from prettytable import PrettyTable
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook
import logger
import json
import os
import sys
import zipfile
import time


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


# region Constants


datetime_format = "%Y-%m-%d %H:%M:%S"
tables_path = resource_path(os.path.join("credentials", "excel_tables.json"))


# endregion


# region Utils


def permute(text: str):
    words = text.split(" ")
    return [" ".join(p) for p in permutations(words)]


def flip(text: str):
    return " ".join(text.split(" ")[::-1])


def similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def find_similar(query: str, strings: list[str]) -> list[str]:
    matches = [s for s in strings if query.lower() in s.lower()]
    if not matches:
        scored_matches = sorted(strings, key=lambda s: similarity(query, s), reverse=True)
        matches = [s for s in scored_matches if similarity(query, s) > 0.5]
    return matches[:5]


def sort_values_by_headers(russian_headers, values, keys_headers):
    header_to_key = swap(keys_headers)
    sorted_keys = [header_to_key[header] for header in russian_headers]
    value_dict = dict(zip(keys_headers.keys(), values))
    sorted_values = [value_dict[key] for key in sorted_keys]
    return sorted_values


def print_table(rows: list[list], headers: list[str]):
    table = PrettyTable()
    table.field_names = headers
    for row in rows:
        table.add_row(row)
    print(table)


def swap(d: dict):
    return {v: k for k, v in d.items()}


def singleton(cls):
    instances = {}

    def getinstance(*args, **kwargs):
        if cls not in instances:
            instances[cls] = cls(*args, **kwargs)
        return instances[cls]

    return getinstance
# endregion


# region Connection


logger = logger.Logger()
tables_data = None
workbook = None
isFirstCreation = False

if workbook is None:
    with open(tables_path, "r", encoding="utf-8") as f:
        tables_data = json.load(f)
    print(f"Opening workbook")

    try:
        # Пытаемся загрузить существующую книгу
        try:
            workbook = load_workbook(tables_data["excel_file_path"])
        except (FileNotFoundError, KeyError, zipfile.BadZipFile):
            # Если файл не существует или поврежден, создаем новую книгу
            workbook = Workbook()
            # Удаляем лист по умолчанию
            for sheet in workbook.sheetnames:
                workbook.remove(workbook[sheet])
            # Создаем необходимые листы
            workbook.create_sheet(tables_data["keys_accounting_wks"])
            workbook.create_sheet(tables_data["keys_wks"])
            workbook.create_sheet(tables_data["employees_wks"])

            isFirstCreation = True

            workbook.save(tables_data["excel_file_path"])
            print(f"Created new workbook at {tables_data['excel_file_path']}")

    except Exception as e:
        print(f"Error while opening/creating workbook: {e}")
        raise

    print(f"Workbook opened")


# endregion


# region Classes


_global_workbook = None
_global_last_reload_time = 0
_global_last_file_mtime = 0


class BaseTable:
    sheet_name: str
    reload_interval = tables_data["excel_reload_interval"]  # в секундах

    def __init__(self):
        self._file_path = tables_data["excel_file_path"]
        self._init_workbook()

    def _init_workbook(self):
        """Инициализация workbook и worksheet"""
        self._check_reload(force=True)

    def _check_reload(self, force=False):
        """Проверяет необходимость перезагрузки общего workbook"""
        global _global_workbook, _global_last_reload_time, _global_last_file_mtime

        now = time.time()
        try:
            current_mtime = os.path.getmtime(self._file_path)

            if (
                force or
                _global_workbook is None or
                now - _global_last_reload_time > self.reload_interval or
                current_mtime > _global_last_file_mtime
            ):
                _global_workbook = load_workbook(self._file_path)
                _global_last_reload_time = now
                _global_last_file_mtime = current_mtime
                # print(f"[BaseTable] Reloaded workbook from file")

            if self.sheet_name not in _global_workbook.sheetnames:
                _global_workbook.create_sheet(self.sheet_name)

            self.wb = _global_workbook
            self.ws = self.wb[self.sheet_name]

        except Exception as err:
            print(f"[BaseTable] Error checking or loading workbook: {err}")
            raise

    def _save_workbook(self):
        """Сохраняет общий workbook в файл"""
        global _global_last_reload_time, _global_last_file_mtime

        try:
            self.wb.save(self._file_path)
            _global_last_reload_time = time.time()
            _global_last_file_mtime = os.path.getmtime(self._file_path)
            # print(f"[BaseTable] Saved workbook to file")
        except Exception as err:
            print(f"[BaseTable] Error saving workbook: {err}")
            raise


class Entry:
    def __init__(
            self,
            key_name: str,
            emp_firstname: str,
            emp_lastname: str,
            emp_phone: str,
            time_received: datetime,
            time_returned: datetime,
            comment: str,
            row: int = None
    ):
        self.key_name = key_name
        self.emp_firstname = emp_firstname
        self.emp_lastname = emp_lastname
        self.emp_phone = emp_phone
        self.comment = comment
        self.row = row

        if isinstance(time_received, str):
            self.time_received = datetime.strptime(time_received, datetime_format)
        elif isinstance(time_received, datetime):
            self.time_received = time_received
        else:
            raise TypeError(
                f"time_received must be a datetime object or a string in '%d.%m.%Y %H:%M:%S' format Current value: {time_received}")

        if isinstance(time_returned, str) and not time_returned.strip() == "":
            self.time_returned = datetime.strptime(time_returned, datetime_format)
        elif isinstance(time_returned, datetime):
            self.time_returned = time_returned
        elif not time_returned:
            self.time_returned = None
        else:
            raise TypeError(
                f"time_returned must be a datetime object or a string in '%d.%m.%Y %H:%M:%S' format. Current value: {time_returned}")

    def __repr__(self):
        return (
            "----------\n"
            f"Key: {self.key_name}\n"
            f"Employee: {self.emp_firstname} {self.emp_lastname} ({self.emp_phone})\n"
            f"Received: {self.time_received.strftime('%d.%m.%Y %H:%M:%S')}\n"
            f"Returned: {self.time_returned.strftime('%d.%m.%Y %H:%M:%S') if self.time_returned else 'Not returned'}\n"
            f"Comment: {self.comment}"
            "\n----------\n"
        )


class KeysAccountingTable(BaseTable):
    def __init__(self):
        self.keys_headers = {
            "key_name": "Ключ",
            "emp_firstname": "Имя",
            "emp_lastname": "Фамилия",
            "emp_phone": "Номер телефона",
            "time_received": "Время получения",
            "time_returned": "Время сдачи",
            "comment": "Комментарий",
        }
        self.sheet_name = tables_data["keys_accounting_wks"]
        super().__init__()

    def new_entry(self, key_name: str, emp_firstname: str, emp_lastname: str, emp_phone: str, comment: str = ""):
        self._check_reload()
        if not comment: comment = ""
        self.append_entry(Entry(key_name, emp_firstname, emp_lastname, emp_phone, datetime.now(), None, comment))

    def setup_table(self):
        print("Setting up keys accounting table")
        if self.ws.max_row == 0 or not any(cell.value for cell in self.ws[1]):
            self.ws.delete_rows(1, self.ws.max_row)
            self.ws.append(list(self.keys_headers.values()))
            self._save_workbook()

    def get_headers(self):
        self._check_reload()
        return [cell.value for cell in self.ws[1]][:len(self.keys_headers)]

    def append_entry(self, entry: Entry):
        self._check_reload()
        print("Appending entry:", entry)
        headers = self.get_headers()
        values = []
        for header in headers:
            key = swap(self.keys_headers)[header]
            val = getattr(entry, key)
            if isinstance(val, datetime):
                val = val.strftime(datetime_format)
            values.append(val)
        self.ws.append(values)
        self._save_workbook()

    def get_all_entries(self) -> list[Entry]:
        self._check_reload()
        rows = list(self.ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = rows[0]
        rows = rows[1:]
        entries = []
        for index, row in enumerate(rows, 2):
            if not any(row):  # Skip empty rows
                continue
            row = [str(x).strip() if x is not None else "" for x in row][:len(self.keys_headers)]
            row = sort_values_by_headers(headers, row, self.keys_headers)
            row.append(index)
            try:
                entries.append(Entry(*row))
            except ValueError as err:
                print(f"Error in row {index}: {row}, {err}")
                pass
        return entries

    def get_not_returned_keys(self) -> list[Entry]:
        self._check_reload()
        entries = self.get_all_entries()
        not_returned_keys = []
        for entry in entries:
            if entry.time_returned is None:
                not_returned_keys.append(entry)
        return not_returned_keys

    def set_return_time(self, entry: Entry, time_returned: datetime = None) -> None:
        self._check_reload()
        headers = self.get_headers()
        if time_returned is None:
            time_returned = datetime.now().strftime(datetime_format)
        col_idx = headers.index(self.keys_headers["time_returned"]) + 1
        self.ws.cell(row=entry.row, column=col_idx, value=time_returned)
        self._save_workbook()

    def set_return_time_by_key_name(self, key_name: str, time_returned: datetime = None) -> None:
        self._check_reload()
        entries = self.get_not_returned_keys()
        for entry in entries:
            if entry.key_name == key_name:
                self.set_return_time(entry, time_returned)
                return


@dataclass
class Key:
    key_name: str
    count: int
    key_type: str
    hardware_type: str


class KeysTable(BaseTable):
    def __init__(self):
        self.keys_headers = {
            "key_name": "Ключ",
            "count": "Количество",
            "key_type": "Тип ключа",
            "hardware_type": "Тип (Аппаратный)",
        }
        self.sheet_name = tables_data["keys_wks"]
        super().__init__()

    def get_by_name(self, name: str) -> Key | None:
        self._check_reload()
        keys = self.get_all_keys()
        for key in keys:
            if key.key_name == name:
                return key
        return None

    def setup_table(self):
        print("Setting up keys table")
        if self.ws.max_row == 0 or not any(cell.value for cell in self.ws[1]):
            self.ws.delete_rows(1, self.ws.max_row)
            self.ws.append(list(self.keys_headers.values()))
            self._save_workbook()

    def get_headers(self):
        self._check_reload()
        return [cell.value for cell in self.ws[1]][:len(self.keys_headers)]

    def new_key(self, key_name, count):
        self._check_reload()
        self.add_key(Key(key_name, count, "None", "None"))

    def add_key(self, key_obj: Key):
        self._check_reload()
        headers = self.get_headers()
        values = []
        for header in headers:
            key = swap(self.keys_headers)[header]
            val = getattr(key_obj, key)
            values.append(val)
        self.ws.append(values)
        self._save_workbook()

    def get_all_keys(self) -> list[Key]:
        self._check_reload()
        rows = list(self.ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = rows[0]
        rows = rows[1:]
        keys = []
        for row in rows:
            if not any(row):  # Skip empty rows
                continue
            row = [str(x).strip() if x is not None else "" for x in row][:len(self.keys_headers)]
            while len(row) < len(self.keys_headers):
                row.append("")
            row = sort_values_by_headers(headers, row, self.keys_headers)
            try:
                keys.append(Key(*row))
            except ValueError:
                print(f"Error in table keys in row {row}")
                pass
        return keys


class Employee:
    def __init__(
            self,
            first_name: str,
            last_name: str,
            phone_number: str,
            telegram: str,
            roles: list[str]) -> None:
        self.first_name = first_name
        self.last_name = last_name
        self.phone_number = phone_number
        self.telegram = telegram

        if isinstance(roles, list):
            self.roles = roles
        elif isinstance(roles, str):
            if roles.strip() == "":
                self.roles = []
            else:
                self.roles = list(map(str.strip, roles.split(", ")))
        else:
            self.roles = []

    def __repr__(self):
        return (
            "----------\n"
            f"Name: {self.first_name} {self.last_name}\n"
            f"Phone number: {self.phone_number}\n"
            f"Telegram: {self.telegram}\n"
            f"Roles: {', '.join(self.roles) if self.roles else "NO ROLES"}"
            "\n----------\n"
        )


class EmployeesTable(BaseTable):
    def __init__(self):
        self.keys_headers = {
            "first_name": "Имя",
            "last_name": "Фамилия",
            "phone_number": "Телефон",
            "telegram": "Телеграм",
            "roles": "Роли",
        }
        self.sheet_name = tables_data["employees_wks"]
        super().__init__()

    def setup_table(self):
        print("Setting up employees table")
        if self.ws.max_row == 0 or not any(cell.value for cell in self.ws[1]):
            self.ws.delete_rows(1, self.ws.max_row)
            self.ws.append(list(self.keys_headers.values()))
            self._save_workbook()

    def get_headers(self):
        self._check_reload()
        return [cell.value for cell in self.ws[1]][:len(self.keys_headers)]

    def new_employee(
            self,
            first_name: str,
            last_name: str,
            phone: str,
            telegram: str,
            roles: list[str] = None
    ):
        self._check_reload()
        self.add_employee(Employee(first_name, last_name, phone, telegram, roles))

    def add_employee(self, employee_obj: Employee):
        self._check_reload()
        headers = self.get_headers()
        values = []
        for header in headers:
            key = swap(self.keys_headers)[header]
            val = getattr(employee_obj, key)
            if isinstance(val, list):
                val = ", ".join(val)
            values.append(str(val))
        self.ws.append(values)
        self._save_workbook()

    def get_all_employees(self) -> list[Employee]:
        self._check_reload()
        rows = list(self.ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = rows[0]
        rows = rows[1:]
        employees = []
        for row in rows:
            if not any(row):  # Skip empty rows
                continue
            row = [str(x).strip() if x is not None else "" for x in row][:len(self.keys_headers)]
            row = sort_values_by_headers(headers, row, self.keys_headers)
            try:
                employees.append(Employee(*row))
            except ValueError:
                print(f"Error in table employees in row {row}")
                pass
        return employees

    def get_by_telegram(self, telegram: str):
        self._check_reload()
        telegram = str(telegram)
        employees = self.get_all_employees()
        for employee in employees:
            if employee.telegram == telegram:
                return employee

    def get_security_employee(self):
        self._check_reload()
        employees = self.get_all_employees()
        for employee in employees:
            if "security" in employee.roles:
                return employee

    def get_by_name(self, first_name: str, last_name: str):
        self._check_reload()
        employees = self.get_all_employees()
        for employee in employees:
            if employee.first_name == first_name and employee.last_name == last_name:
                return employee


if isFirstCreation:
    kat = KeysAccountingTable()
    kat.setup_table()
    del kat
    keys = KeysTable()
    keys.setup_table()
    del keys
    emps = EmployeesTable()
    emps.setup_table()
    del emps


# endregion


# region Tests


'''def main():
    """Test code"""
    # Инициализация таблиц
    kat = KeysAccountingTable()
    keys = KeysTable()
    emps = EmployeesTable()

    kat.setup_table()
    keys.setup_table()
    emps.setup_table()

    # Создаем тестовые данные
    test_key = "Al1234"
    test_firstname = "Иван"
    test_lastname = "Тестов"
    test_phone = "+79991112233"
    test_telegram = "@testuser"

    try:
        # Тестирование таблицы ключей
        print("\n=== Тестирование таблицы ключей ===")
        keys.new_key(test_key, 5)
        all_keys = keys.get_all_keys()
        print(f"Все ключи ({len(all_keys)}):")
        for key in all_keys:
            print(f"- {key.key_name}: {key.count} шт.")

        # Тестирование таблицы сотрудников
        print("\n=== Тестирование таблицы сотрудников ===")
        emps.new_employee(test_firstname, test_lastname, test_phone, test_telegram, ["tester"])
        all_employees = emps.get_all_employees()
        print(f"Все сотрудники ({len(all_employees)}):")
        for emp in all_employees:
            print(f"- {emp.first_name} {emp.last_name} ({emp.telegram}): {', '.join(emp.roles)}")

        # Тестирование таблицы учета
        print("\n=== Тестирование таблицы учета ключей ===")
        kat.new_entry(test_key, test_firstname, test_lastname, test_phone, "тестовая выдача")
        all_entries = kat.get_all_entries()
        print(f"Все записи учета ({len(all_entries)}):")
        for entry in all_entries:
            print(f"- Ключ: {entry.key_name}, Сотрудник: {entry.emp_firstname} {entry.emp_lastname}")

        # Тестирование возврата ключа
        print("\n=== Тестирование возврата ключа ===")
        not_returned = kat.get_not_returned_keys()
        print(f"Не возвращенные ключи ({len(not_returned)}):")
        for entry in not_returned:
            print(f"- {entry.key_name} (строка {entry.row})")
            # Возвращаем первый найденный ключ
            kat.set_return_time(entry)
            print(f"Ключ {entry.key_name} возвращен")
            break

        # Проверка после возврата
        not_returned_after = kat.get_not_returned_keys()
        print(f"Не возвращенные ключи после возврата ({len(not_returned_after)})")

        # Поиск сотрудника
        print("\n=== Тестирование поиска ===")
        found_emp = emps.get_by_telegram(test_telegram)
        if found_emp:
            print(f"Найден сотрудник по Telegram: {found_emp.first_name} {found_emp.last_name}")

        found_key = keys.get_by_name(test_key)
        if found_key:
            print(f"Найден ключ: {found_key.key_name} (кол-во: {found_key.count})")

        print("\nТестирование завершено успешно!")

    # except Exception as e:
        # print(f"\nОшибка при тестировании: {e}")
    finally:
        # Сохраняем изменения
        workbook.save(tables_data["excel_file_path"])
        print(f"\nДанные сохранены в файл: {tables_data['excel_file_path']}")


if __name__ == "__main__":
    main()'''


# endregion
